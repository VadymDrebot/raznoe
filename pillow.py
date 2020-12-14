from openpyxl import load_workbook
data_file = 'polit.xlsx'
wb = load_workbook(data_file)
print("Found the following worksheets:")
for sheetname in wb.sheetnames:   # для каждого листа из книги...
    print(sheetname)              # печать каждого листа
a=[]
ws = wb['students_excel']         # ws= лист "students_excel" из книги "polit"
all_rows = list(ws.rows)
b=len(all_rows)                   # b=6 --колличество СТОЛБЦОВ с строке--длина списка LIST
for i in range(1,b):              # чередуем строки из листа
    for cell in all_rows[i]:      # чередуем  ячейки каждой строки
        a.append(cell.value)      # достаем ЗНАЧЕНИЕ каждлй ячейки и вставляем в список
    print(a)
    a=[]




